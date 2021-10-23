import equilibrafases as EQ
import numpy as npy
import openpyxl as wb
import os
import array as ar

## matriz armazenando o numero de fases que o circuito demanda e a potencia demandada
print ("Esse aplicativo foi idealizado e implementado como atividade de Iniciaçao Científica pelo aluno:")
print ("\t Paulo Andre de Toledo - Engenharia Civil - Universidade Federal Fluminense")
print(" Voce e livre para usar o aplicativo para fins academicos devendo citar a fonte e seus criadores.")
print(" Voce não pode usar o aplicativo, seu codigo fonte, algoritmos ou qualquer conceito incorporado para fins comerciais.")
print(" Voce deve informar o nome de um arquivo xlsx do Excel que contenha uma planilha chamada eqf_circuitos")
print(" A celula A1 da planilha eqf_circuitos deve conter o numero total de circuitos do seu quadro (nc)")
print(" As celulas A2:Bnc devem conter na coluna A o numero de fases do circuto (1,2,3) ")
print(" As celulas A2:Bnc devem conter na coluna B a potencia designada para o circuito ")
print(" Um quadro otimizado sera criado e salvo na planilha chamada eqf_quadro")
#nomeplan = (input('Informe o nome do arquivo Excel a ser processado: ') )
nomeplan=input("Entre com o nome e caminho do arquivo: ")
while os.path.exists(nomeplan)==False:
    nomeplan=input("Entre com o nome e caminho do arquivo: ")

ws = wb.load_workbook(nomeplan)
planin = ws['eqf_circuitos']
NC = int(planin.cell(1, 1).value)

Quadroxls = npy.zeros([NC,2],dtype=float)
planout = ws.create_sheet('eqf_quadro')

for a in range(2,NC+2):

    Quadroxls[a-2,0] = planin.cell(a,1).value
    Quadroxls[a - 2, 1] = planin.cell(a, 2).value


pop = EQ.geraPopulacao(20, Quadroxls)


sol = EQ.equilibraFases(pop, Quadroxls,750)

I = EQ.funcaoObjetivo(sol)

print("Quadro otimizado:\n",sol)

soma = sol.sum(0)
print("Divisão por fases: \n",soma)
soma /= sol.sum()

print("Divisão percentual por fase:\n",soma*100)

print("Diferença percentual entre as fases com maior carga e a com menor carga:\n",(max(soma)- min(soma))*100)

lin = 1
col = 1

for a in range(len(sol)):
    for c in range(0, 3):
       planout.cell(lin+a, col+c).value = sol[a][c]


ws.save(nomeplan)
ws.close()







